# module_app/utils/excel_export.py
"""
Экспорт и импорт данных в Excel
"""

import io
import re
from datetime import datetime
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter
from django.db import transaction
from django.core.exceptions import ValidationError
from ..models import EmployeeProfile, WorkSchedule


def export_employees_to_excel(employees):
    """
    Экспорт сотрудников в Excel

    Args:
        employees: QuerySet сотрудников

    Returns:
        Workbook: объект Excel
    """
    wb = Workbook()
    ws = wb.active
    ws.title = 'Сотрудники'

    # Заголовки
    headers = [
        'Табельный номер', 'ФИО', 'Должность', 'Отдел', 'Телефон',
        'Email', 'Дата рождения', 'Дата приема', 'Дата увольнения',
        'Тип занятости', 'Оклад', 'ИНН', 'СНИЛС', 'Статус'
    ]

    # Стили заголовков
    header_font = Font(bold=True, size=11, color='FFFFFF')
    header_fill = PatternFill(start_color='4e73df', end_color='4e73df', fill_type='solid')
    header_alignment = Alignment(horizontal='center', vertical='center')

    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment

    # Данные
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    for row_idx, emp in enumerate(employees, 2):
        data = [
            emp.employee_code or '',
            emp.full_name,
            emp.position or '',
            emp.department or '',
            emp.phone or '',
            emp.user.email,
            emp.birth_date.strftime('%d.%m.%Y') if emp.birth_date else '',
            emp.hire_date.strftime('%d.%m.%Y') if emp.hire_date else '',
            emp.dismissal_date.strftime('%d.%m.%Y') if emp.dismissal_date else '',
            dict(EmployeeProfile.EMPLOYMENT_TYPES).get(emp.employment_type, ''),
            float(emp.salary) if emp.salary else '',
            emp.tax_id or '',
            emp.snils or '',
            'Активен' if emp.is_active else 'Уволен'
        ]

        for col, value in enumerate(data, 1):
            cell = ws.cell(row=row_idx, column=col, value=value)
            cell.border = thin_border
            cell.alignment = Alignment(horizontal='left', vertical='center')

    # Автоширина колонок
    for col in range(1, len(headers) + 1):
        column_letter = get_column_letter(col)
        max_length = 0
        for row in range(1, ws.max_row + 1):
            cell_value = ws.cell(row=row, column=col).value
            if cell_value:
                max_length = max(max_length, len(str(cell_value)))
        ws.column_dimensions[column_letter].width = min(max_length + 2, 50)

    return wb


def import_employees_from_excel(file):
    """
    Импорт сотрудников из Excel

    Args:
        file: загруженный файл Excel

    Returns:
        dict: результаты импорта (добавлено, обновлено, ошибки)
    """
    wb = load_workbook(file)
    ws = wb.active

    results = {
        'added': 0,
        'updated': 0,
        'errors': []
    }

    # Пропускаем заголовок
    for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), 2):
        if not row[1]:  # ФИО обязательно
            continue

        try:
            # Поиск сотрудника по табельному номеру или ФИО
            employee_code = row[0]
            full_name = row[1]

            employee = None
            if employee_code:
                employee = EmployeeProfile.objects.filter(employee_code=employee_code).first()

            if not employee:
                # Ищем по ФИО
                name_parts = full_name.split()
                if len(name_parts) >= 2:
                    employee = EmployeeProfile.objects.filter(
                        user__first_name=name_parts[0],
                        user__last_name=name_parts[-1]
                    ).first()

            with transaction.atomic():
                if employee:
                    # Обновление существующего
                    employee.position = row[2] or employee.position
                    employee.department = row[3] or employee.department
                    employee.phone = row[4] or employee.phone

                    if row[6]:  # Дата рождения
                        try:
                            employee.birth_date = datetime.strptime(row[6], '%d.%m.%Y').date()
                        except:
                            pass

                    if row[7]:  # Дата приема
                        try:
                            employee.hire_date = datetime.strptime(row[7], '%d.%m.%Y').date()
                        except:
                            pass

                    employee.tax_id = row[11] or employee.tax_id
                    employee.snils = row[12] or employee.snils
                    employee.save()

                    results['updated'] += 1
                else:
                    # Создание нового
                    # Требуется также создать пользователя
                    from django.contrib.auth.models import User

                    name_parts = full_name.split()
                    first_name = name_parts[0] if len(name_parts) >= 1 else ''
                    last_name = name_parts[-1] if len(name_parts) >= 2 else ''

                    username = re.sub(r'[^a-zA-Z0-9]', '', f"{first_name}{last_name}_{datetime.now().timestamp()}")

                    user = User.objects.create_user(
                        username=username[:150],
                        password='change_me',
                        first_name=first_name,
                        last_name=last_name,
                        email=row[5] or f"{username}@temp.com"
                    )

                    employee = EmployeeProfile.objects.create(
                        user=user,
                        employee_code=employee_code,
                        position=row[2] or '',
                        department=row[3] or '',
                        phone=row[4] or '',
                        tax_id=row[11] or '',
                        snils=row[12] or ''
                    )

                    if row[6]:  # Дата рождения
                        try:
                            employee.birth_date = datetime.strptime(row[6], '%d.%m.%Y').date()
                            employee.save()
                        except:
                            pass

                    if row[7]:  # Дата приема
                        try:
                            employee.hire_date = datetime.strptime(row[7], '%d.%m.%Y').date()
                            employee.save()
                        except:
                            pass

                    results['added'] += 1

        except Exception as e:
            results['errors'].append({
                'row': row_idx,
                'data': row,
                'error': str(e)
            })

    return results


def export_attendance_to_excel(attendance_logs):
    """
    Экспорт табеля в Excel

    Args:
        attendance_logs: QuerySet записей табеля

    Returns:
        Workbook: объект Excel
    """
    wb = Workbook()
    ws = wb.active
    ws.title = 'Табель учета'

    headers = ['Дата', 'Сотрудник', 'Должность', 'Отдел', 'Время', 'Событие', 'Часы']

    header_font = Font(bold=True, size=11, color='FFFFFF')
    header_fill = PatternFill(start_color='4e73df', end_color='4e73df', fill_type='solid')

    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center')

    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    for row_idx, log in enumerate(attendance_logs, 2):
        data = [
            log.date.strftime('%d.%m.%Y'),
            log.employee.full_name,
            log.employee.position or '',
            log.employee.department or '',
            log.time.strftime('%H:%M'),
            log.get_event_display(),
            log.hours if log.hours else ''
        ]

        for col, value in enumerate(data, 1):
            cell = ws.cell(row=row_idx, column=col, value=value)
            cell.border = thin_border
            cell.alignment = Alignment(horizontal='left')

    for col in range(1, len(headers) + 1):
        ws.column_dimensions[get_column_letter(col)].width = 15

    return wb


def export_to_excel(data, filename):
    """
    Универсальный экспорт в Excel (обертка)

    Args:
        data: список словарей с данными
        filename: имя файла
    """
    from django.http import HttpResponse
    from openpyxl import Workbook

    wb = Workbook()
    ws = wb.active
    ws.title = 'Данные'

    if data:
        # Заголовки
        headers = list(data[0].keys())
        for col, header in enumerate(headers, 1):
            ws.cell(row=1, column=col, value=header)

        # Данные
        for row_idx, row_data in enumerate(data, 2):
            for col_idx, key in enumerate(headers, 1):
                ws.cell(row=row_idx, column=col_idx, value=str(row_data.get(key, '')))

    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename={filename}.xlsx'
    wb.save(response)
    return response


def import_from_excel(file):
    """
    Универсальный импорт из Excel (обертка)
    """
    return import_employees_from_excel(file)