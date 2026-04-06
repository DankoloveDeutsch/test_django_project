# module_app/utils/report_generator.py
"""
Генерация отчетов (Excel, PDF, CSV)
"""

import io
import csv
from datetime import datetime, timedelta
from django.db.models import Sum, Count, Avg
from django.utils import timezone
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4, landscape
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import cm
from reportlab.lib.enums import TA_CENTER, TA_LEFT
from reportlab.pdfbase import pdfmetrics
from reportlab.pdfbase.ttfonts import TTFont
from ..models import EmployeeProfile, AttendanceLog, MonthlyReport, Reminder


def export_to_excel(data, sheet_name='Report', filename=None):
    """
    Экспорт данных в Excel

    Args:
        data: список словарей с данными
        sheet_name: название листа
        filename: имя файла (опционально)

    Returns:
        HttpResponse: ответ с файлом Excel
    """
    from django.http import HttpResponse

    wb = Workbook()
    ws = wb.active
    ws.title = sheet_name

    if not data:
        ws['A1'] = 'Нет данных'
        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = f'attachment; filename={filename or "report"}.xlsx'
        wb.save(response)
        return response

    # Заголовки
    headers = list(data[0].keys())
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = Font(bold=True, size=12)
        cell.alignment = Alignment(horizontal='center')
        cell.fill = PatternFill(start_color='CCCCCC', end_color='CCCCCC', fill_type='solid')

    # Данные
    for row_idx, row_data in enumerate(data, 2):
        for col_idx, key in enumerate(headers, 1):
            value = row_data.get(key, '')
            ws.cell(row=row_idx, column=col_idx, value=value)

    # Автоширина колонок
    for col in range(1, len(headers) + 1):
        column_letter = get_column_letter(col)
        max_length = max(len(str(ws.cell(row=row, column=col).value or ''))
                         for row in range(1, ws.max_row + 1))
        ws.column_dimensions[column_letter].width = min(max_length + 2, 50)

    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename={filename or "report"}.xlsx'
    wb.save(response)
    return response


def export_to_csv(data, filename=None):
    """
    Экспорт данных в CSV

    Args:
        data: список словарей с данными
        filename: имя файла

    Returns:
        HttpResponse: ответ с файлом CSV
    """
    from django.http import HttpResponse

    if not data:
        data = [{'Сообщение': 'Нет данных'}]

    headers = list(data[0].keys())

    response = HttpResponse(content_type='text/csv; charset=utf-8')
    response['Content-Disposition'] = f'attachment; filename={filename or "report"}.csv'
    response.write('\ufeff')  # BOM для UTF-8

    writer = csv.writer(response)
    writer.writerow(headers)

    for row in data:
        writer.writerow([row.get(h, '') for h in headers])

    return response


def generate_monthly_report_pdf(employee, year, month):
    """
    Генерация месячного отчета в PDF

    Args:
        employee: объект EmployeeProfile
        year: год
        month: месяц

    Returns:
        bytes: PDF контент
    """
    buffer = io.BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=A4)

    styles = getSampleStyleSheet()
    styles.add(ParagraphStyle(name='CenterTitle', alignment=TA_CENTER, fontSize=16, leading=20, spaceAfter=20))
    styles.add(ParagraphStyle(name='Right', alignment=TA_RIGHT))

    story = []

    # Заголовок
    title = f"ТАБЕЛЬ УЧЕТА РАБОЧЕГО ВРЕМЕНИ\n{employee.full_name}\n{month:02d}.{year}"
    story.append(Paragraph(title, styles['CenterTitle']))
    story.append(Spacer(1, 0.5 * cm))

    # Данные за месяц
    logs = AttendanceLog.objects.filter(
        employee=employee,
        date__year=year,
        date__month=month
    ).order_by('date')

    # Таблица
    data = [['Дата', 'Приход', 'Уход', 'Обед', 'Часы']]

    for log in logs:
        data.append([
            log.date.strftime('%d.%m.%Y'),
            log.time.strftime('%H:%M') if log.event == 'start' else '',
            log.time.strftime('%H:%M') if log.event == 'end' else '',
            '',
            str(log.hours) if log.hours else ''
        ])

    # Итоги
    total_hours = logs.aggregate(Sum('hours'))['hours__sum'] or 0
    data.append(['', '', '', 'ИТОГО:', f'{total_hours:.2f}'])

    table = Table(data, colWidths=[2.5 * cm, 2.5 * cm, 2.5 * cm, 2.5 * cm, 2.5 * cm])
    table.setStyle(TableStyle([
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('GRID', (0, 0), (-1, -1), 0.5, colors.black),
        ('FONTNAME', (0, -1), (-1, -1), 'Helvetica-Bold'),
    ]))

    story.append(table)

    doc.build(story)
    buffer.seek(0)
    return buffer.getvalue()


def generate_government_report_pdf(report_type, period, data):
    """
    Генерация отчета для госорганов

    Args:
        report_type: тип отчета (pension_fund/tax_service/social_fund)
        period: период
        data: данные для отчета

    Returns:
        bytes: PDF контент
    """
    buffer = io.BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=A4)

    styles = getSampleStyleSheet()
    styles.add(ParagraphStyle(name='CenterTitle', alignment=TA_CENTER, fontSize=14, leading=18))

    story = []

    # Заголовок
    titles = {
        'pension_fund': 'Отчет в Пенсионный фонд РФ (СФР)',
        'tax_service': 'Отчет в Федеральную налоговую службу',
        'social_fund': 'Отчет в Фонд социального страхования'
    }

    story.append(Paragraph(titles.get(report_type, 'Отчет'), styles['CenterTitle']))
    story.append(Paragraph(f"Период: {period}", styles['Normal']))
    story.append(Spacer(1, 0.5 * cm))

    # Таблица с данными
    if data:
        headers = list(data[0].keys()) if data else ['Показатель', 'Значение']
        table_data = [headers]

        for row in data:
            table_data.append([str(row.get(h, '')) for h in headers])

        table = Table(table_data)
        table.setStyle(TableStyle([
            ('GRID', (0, 0), (-1, -1), 0.5, colors.black),
            ('BACKGROUND', (0, 0), (-1, 0), colors.lightgrey),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ]))

        story.append(table)

    doc.build(story)
    buffer.seek(0)
    return buffer.getvalue()


def generate_department_report_pdf(department, year, quarter=None):
    """
    Генерация отчета по отделу

    Args:
        department: название отдела
        year: год
        quarter: квартал (опционально)

    Returns:
        bytes: PDF контент
    """
    employees = EmployeeProfile.objects.filter(department=department, is_active=True)

    buffer = io.BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=landscape(A4))

    styles = getSampleStyleSheet()
    styles.add(ParagraphStyle(name='CenterTitle', alignment=TA_CENTER, fontSize=14, leading=18))

    story = []

    title = f"ОТЧЕТ ПО ОТДЕЛУ: {department or 'Без отдела'}\nЗа {year} год"
    if quarter:
        title += f", {quarter} квартал"
    story.append(Paragraph(title, styles['CenterTitle']))
    story.append(Spacer(1, 0.5 * cm))

    # Данные по сотрудникам
    data = [['Табельный номер', 'ФИО', 'Должность', 'Часы', 'Переработки', 'Документы']]

    for emp in employees:
        # Часы за период
        if quarter:
            months = range((quarter - 1) * 3 + 1, quarter * 3 + 1)
            total_hours = 0
            for month in months:
                month_hours = AttendanceLog.objects.filter(
                    employee=emp, date__year=year, date__month=month
                ).aggregate(Sum('hours'))['hours__sum'] or 0
                total_hours += month_hours
        else:
            total_hours = MonthlyReport.objects.filter(employee=emp).aggregate(Sum('total_hours'))[
                              'total_hours__sum'] or 0

        overtime = max(0, total_hours - (len(employees) * 160 * (3 if quarter else 12)))

        data.append([
            emp.employee_code or '',
            emp.full_name,
            emp.position or '',
            f'{total_hours:.1f}',
            f'{overtime:.1f}',
            str(emp.personal_documents.count())
        ])

    table = Table(data, colWidths=[2.5 * cm, 4 * cm, 3.5 * cm, 2.5 * cm, 2.5 * cm, 2.5 * cm])
    table.setStyle(TableStyle([
        ('GRID', (0, 0), (-1, -1), 0.5, colors.black),
        ('BACKGROUND', (0, 0), (-1, 0), colors.lightgrey),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
    ]))

    story.append(table)

    doc.build(story)
    buffer.seek(0)
    return buffer.getvalue()


def generate_annual_report(year, department=None):
    """
    Генерация годового отчета

    Args:
        year: год
        department: отдел (опционально)

    Returns:
        dict: данные отчета
    """
    from ..services.report_service import ReportService
    return ReportService.get_yearly_attendance_report(year, department)


def generate_government_report_pdf(report_type, period, data):
    """
    Генерация отчета для госорганов в PDF

    Args:
        report_type: тип отчета
        period: период
        data: данные

    Returns:
        bytes: PDF контент
    """
    from .pdf_export import generate_government_report_pdf as pdf_func
    return pdf_func(report_type, period, data)


def generate_department_report_pdf(department, year, quarter=None):
    """
    Генерация отчета по отделу в PDF

    Args:
        department: отдел
        year: год
        quarter: квартал (опционально)

    Returns:
        bytes: PDF контент
    """
    from .pdf_export import generate_department_report_pdf as pdf_func
    return pdf_func(department, year, quarter)